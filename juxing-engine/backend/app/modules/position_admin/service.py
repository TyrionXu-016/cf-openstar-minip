from __future__ import annotations

import csv
import io
import json

from openpyxl import load_workbook
from sqlalchemy import delete, func, or_, select
from sqlalchemy.orm import Session

from app.db.models import FavoritePosition, Position
from app.modules.admin_import_undo.service import RESOURCE_POSITIONS, record_import_batch
from app.modules.position_admin.schemas import ImportResult, PositionInput, PositionItem


def upsert_one(db: Session, payload: PositionInput, on_conflict: str = "upsert") -> tuple[str, Position | None]:
    row = db.get(Position, payload.id)
    full_payload = _build_payload(payload)
    if row:
        if on_conflict == "skip":
            return "skipped", None
        row.name = payload.name.strip()
        row.category = payload.category.strip()
        row.sub_category = (payload.subCategory or "").strip() or None
        row.description = (payload.description or "").strip() or None
        row.is_three_free = payload.isThreeFree
        row.payload = json.dumps(full_payload, ensure_ascii=False)
        return "updated", row

    row = Position(
        id=payload.id,
        name=payload.name.strip(),
        category=payload.category.strip(),
        sub_category=(payload.subCategory or "").strip() or None,
        description=(payload.description or "").strip() or None,
        is_three_free=payload.isThreeFree,
        payload=json.dumps(full_payload, ensure_ascii=False),
    )
    db.add(row)
    return "created", row


def import_positions(db: Session, items: list[PositionInput], on_conflict: str = "upsert") -> ImportResult:
    created = 0
    updated = 0
    skipped = 0
    errors: list[str] = []
    created_ids: list[int] = []

    for idx, item in enumerate(items, start=1):
        try:
            action, row = upsert_one(db, item, on_conflict=on_conflict)
            if action == "created":
                created += 1
                if row is not None:
                    created_ids.append(row.id)
            elif action == "updated":
                updated += 1
            else:
                skipped += 1
        except Exception as exc:
            errors.append(f"第{idx}条导入失败: {exc}")

    record_import_batch(db, resource=RESOURCE_POSITIONS, created_ids=created_ids)
    db.commit()
    return ImportResult(
        total=len(items),
        created=created,
        updated=updated,
        skipped=skipped,
        errors=errors,
        created_ids=created_ids,
    )


def list_positions(
    db: Session,
    page: int,
    page_size: int,
    category: str | None = None,
    keyword: str | None = None,
) -> tuple[list[PositionItem], int]:
    stmt = select(Position)
    count_stmt = select(func.count(Position.id))
    filters = []

    if category:
        filters.append(Position.category == category)
    if keyword:
        kw = f"%{keyword}%"
        filters.append(
            or_(Position.name.ilike(kw), Position.sub_category.ilike(kw), Position.description.ilike(kw))
        )

    if filters:
        stmt = stmt.where(*filters)
        count_stmt = count_stmt.where(*filters)

    total = db.scalar(count_stmt) or 0
    rows = db.scalars(stmt.offset((page - 1) * page_size).limit(page_size)).all()
    return [_to_item(row) for row in rows], total


def get_position(db: Session, position_id: int) -> PositionItem | None:
    row = db.get(Position, position_id)
    if not row:
        return None
    return _to_item(row)


def delete_position_by_id(db: Session, position_id: int) -> bool:
    row = db.get(Position, position_id)
    if not row:
        return False
    db.execute(delete(FavoritePosition).where(FavoritePosition.position_id == position_id))
    db.delete(row)
    return True


def parse_csv_bytes(content: bytes) -> list[PositionInput]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return _rows_to_inputs(reader)


def parse_xlsx_bytes(content: bytes) -> list[PositionInput]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []
    normalized_headers = [str(h).strip() if h is not None else "" for h in headers]
    mapped_rows = []
    for row in rows:
        item = {}
        for idx, value in enumerate(row):
            if idx < len(normalized_headers):
                item[normalized_headers[idx]] = value
        mapped_rows.append(item)
    return _rows_to_inputs(mapped_rows)


def _rows_to_inputs(rows: list | csv.DictReader) -> list[PositionInput]:
    items: list[PositionInput] = []
    for idx, raw in enumerate(rows, start=2):
        try:
            row = {str(k).strip(): v for k, v in dict(raw).items()}
            pid = _to_int(row.get("id"))
            name = str(row.get("name") or "").strip()
            if not pid or not name:
                raise ValueError("id 或 name 为空")
            category = str(row.get("category") or "国考").strip()
            sub_category = _to_optional_str(row.get("subCategory"))
            description = _to_optional_str(row.get("description"))
            is_three_free = _to_bool(row.get("isThreeFree"))

            payload_raw = row.get("payload")
            payload = None
            if payload_raw not in (None, ""):
                payload = json.loads(str(payload_raw))

            items.append(
                PositionInput(
                    id=pid,
                    name=name,
                    category=category,
                    subCategory=sub_category,
                    description=description,
                    isThreeFree=is_three_free,
                    payload=payload,
                )
            )
        except Exception as exc:
            raise ValueError(f"第{idx}行解析失败: {exc}") from exc
    return items


def _to_item(row: Position) -> PositionItem:
    payload = {}
    if row.payload:
        try:
            payload = json.loads(row.payload)
        except Exception:
            payload = {}
    payload = payload if isinstance(payload, dict) else {}
    payload["id"] = row.id
    payload["name"] = row.name
    payload["category"] = row.category
    payload["subCategory"] = row.sub_category
    payload["description"] = row.description
    payload["isThreeFree"] = row.is_three_free
    return PositionItem(
        id=row.id,
        name=row.name,
        category=row.category,
        subCategory=row.sub_category,
        description=row.description,
        isThreeFree=row.is_three_free,
        payload=payload,
    )


def _build_payload(payload: PositionInput) -> dict:
    base = payload.payload.copy() if isinstance(payload.payload, dict) else {}
    base["id"] = payload.id
    base["name"] = payload.name.strip()
    base["category"] = payload.category.strip()
    base["subCategory"] = (payload.subCategory or "").strip() or None
    base["description"] = (payload.description or "").strip() or None
    base["isThreeFree"] = payload.isThreeFree
    return base


def _to_optional_str(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _to_int(value: object) -> int:
    if value is None or str(value).strip() == "":
        return 0
    return int(float(str(value).strip()))


def _to_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    return text in {"1", "true", "yes", "y", "是"}

