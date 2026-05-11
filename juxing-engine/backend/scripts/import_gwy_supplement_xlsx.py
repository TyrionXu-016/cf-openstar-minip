#!/usr/bin/env python3
"""将「中央机关及其直属机构年度补充录用公务员职位表」类 Excel 导入岗位库。

表特征（常见）：第 1 行为提示语，第 2 行为列标题（含 职位代码、部门名称、招考职位 等）。

用法:

  cd juxing-engine/backend
  source .venv/bin/activate
  python scripts/import_gwy_supplement_xlsx.py "/path/to/职位表.xlsx"

依赖 backend/.env 中的 SQLITE_PATH；默认 upsert（同 id 会更新）。
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from app.modules.position_admin.schemas import PositionInput
from app.modules.position_admin.service import import_positions
from app.db.session import SessionLocal


def _cell_str(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return str(v)
    return str(v).strip()


def _parse_code(raw: object) -> int:
    s = _cell_str(raw).replace(" ", "")
    if not s:
        raise ValueError("职位代码为空")
    return int(float(s)) if "." in s else int(s)


def _short_category(exam_type: str) -> str:
    """category 字段最长 20，此处压缩考试类别或退回国考。"""
    t = exam_type.strip()
    if not t:
        return "国考"
    if len(t) <= 20:
        return t
    # 常见前缀过长时兜底
    if "行政执法类" in t:
        return "行政执法类"
    if "地市" in t and "综合管理" in t:
        return "地市综合管理类"
    if "省级直属机构综合管理" in t:
        return "省级直属综合管理类"
    return t[:20]


def _is_three_free(remark: str, intro: str, major: str) -> bool:
    blob = f"{remark}{intro}{major}"
    return "三不限" in blob or ("不限" in blob and "专业" in blob and major.count("不限") >= 1)


def _row_to_input(headers: list[str], row: tuple[object, ...]) -> PositionInput:
    d = {}
    for i, h in enumerate(headers):
        key = _cell_str(h)
        if not key:
            continue
        d[key] = row[i] if i < len(row) else None

    pid = _parse_code(d.get("职位代码"))
    dept = _cell_str(d.get("部门名称"))
    title = _cell_str(d.get("招考职位"))
    name = f"{dept}-{title}".strip("-") if dept or title else ""
    if not name:
        name = title or dept or f"职位{pid}"
    name = name[:255]

    exam = _cell_str(d.get("考试类别"))
    category = _short_category(exam)
    sub = _cell_str(d.get("用人司局"))[:255] or None

    intro = _cell_str(d.get("职位简介"))
    major = _cell_str(d.get("专业"))
    edu = _cell_str(d.get("学历"))
    deg = _cell_str(d.get("学位"))
    loc = _cell_str(d.get("工作地点"))
    remark = _cell_str(d.get("备注"))
    interview = _cell_str(d.get("是否在面试阶段组织专业能力测试"))
    ratio = _cell_str(d.get("面试人员比例"))

    desc_parts = [
        f"职位简介：{intro}" if intro else "",
        f"专业：{major}" if major else "",
        f"学历：{edu}" if edu else "",
        f"学位：{deg}" if deg else "",
        f"工作地点：{loc}" if loc else "",
        f"面试专业测试：{interview}；比例：{ratio}" if interview or ratio else "",
        f"备注：{remark}" if remark else "",
    ]
    description = "\n".join(p for p in desc_parts if p) or None

    payload: dict = {}
    for k, v in d.items():
        if v is None or v == "":
            payload[k] = ""
        elif isinstance(v, float) and v == int(v):
            payload[k] = int(v)
        else:
            payload[k] = _cell_str(v)

    return PositionInput(
        id=pid,
        name=name,
        category=category,
        subCategory=sub,
        description=description,
        isThreeFree=_is_three_free(remark, intro, major),
        payload=payload,
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path", type=Path, help="职位表 .xlsx 路径")
    parser.add_argument("--dry-run", action="store_true", help="只解析条数，不写库")
    args = parser.parse_args()

    path = args.xlsx_path.expanduser().resolve()
    if not path.is_file():
        print(f"文件不存在: {path}")
        sys.exit(1)

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    first = next(rows_iter, None)
    second = next(rows_iter, None)
    # 若第一行不像表头（单列长文本），则跳过；否则从第一行起当作表头
    headers_row = second if second and len([x for x in second if x not in (None, "")]) > 5 else first
    if headers_row is second:
        data_rows = rows_iter
    else:
        headers_row = first
        data_rows = rows_iter

    headers = [_cell_str(h) for h in headers_row]
    if "职位代码" not in headers:
        print("未找到列「职位代码」，请确认第二行是否为表头。", headers[:15])
        sys.exit(1)

    raw_items: list[PositionInput] = []
    errors: list[str] = []
    line_no = 3 if headers_row is second else 2
    for row in data_rows:
        line_no += 1
        if not row or not any(x not in (None, "") for x in row):
            continue
        try:
            raw_items.append(_row_to_input(headers, row))
        except Exception as exc:
            errors.append(f"第{line_no}行: {exc}")

    # 官方表可能出现同一「职位代码」多行，按 id 保留最后一次写入
    by_id: dict[int, PositionInput] = {}
    for it in raw_items:
        by_id[it.id] = it
    items = list(by_id.values())
    merged = len(raw_items) - len(items)
    if merged:
        print(f"职位代码重复：已合并 {merged} 行（保留同代码的最后一行）。")

    print(f"解析成功 {len(raw_items)} 行，去重后 {len(items)} 条，失败 {len(errors)} 条。")
    for e in errors[:20]:
        print(" ", e)
    if len(errors) > 20:
        print(f" ... 另有 {len(errors) - 20} 条错误")

    if args.dry_run:
        print("[dry-run] 未写入数据库")
        return

    with SessionLocal() as db:
        result = import_positions(db, items, on_conflict="upsert")

    print(
        f"导入完成: 总计 {result.total}，新建 {result.created}，更新 {result.updated}，跳过 {result.skipped}",
    )
    for e in result.errors[:15]:
        print(" ", e)


if __name__ == "__main__":
    main()
